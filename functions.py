from urllib.request import Request, urlopen, urlretrieve
from urllib.error import URLError, HTTPError
from bs4 import BeautifulSoup
import pandas as pd
import json
import requests
import io
from sheet2dict import Worksheet
import os, sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles.alignment import Alignment
import re

def buscarLigas():
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.100 Safari/537.36'}

    try:    
        
        response = requests.get("https://www.soccerstats.com/js/summaries_leaguesearchjs.js")

        resRaw = response.text.replace("var options=", "").replace("value:", "'Liga':").replace("data:", "'Link':")
        resRaw = resRaw.replace(";$('#autocomplete').autocomplete({lookup:options,onSelect:function(suggestion){}});$('#autocomplete').autocomplete({lookup:options,onSelect:function(suggestion){location.href=suggestion.data}});", "")
        resRaw = resRaw.replace("'", '"')
        json_ligas = json.loads(resRaw)



        listLigas = []
        urlLigas = []
        # print(json.dumps(json_ligas, indent=2))

        for i, l in enumerate(json_ligas):
            listLigas.append(json_ligas[i]['Liga'])
            urlLigas.append(json_ligas[i]['Link'])


        return listLigas, urlLigas
    
    except HTTPError as e:
        print(e.status, e.reason)

    except URLError as e:
        print(e.reason)


def buscarTimeLigas(urlLiga):
    try:
        response = requests.get("https://www.soccerstats.com/"+urlLiga)

        soup = BeautifulSoup(response.content, 'html.parser')

        Times = []
        urlTimes = []
        media_gols_liga = []

        for item in soup.find_all("td"):
            if not "Home goals per match" in item.text:continue
            if(item.font.b != None):
                media_gols_liga.append(item.font.b.text)
                break

        

        Times1 = soup.find('font', string='Teams')
        Times2 = Times1.parent.parent.find_all('a', href=True)





        for i, t in enumerate(Times2):
            Times.append(Times2[i].get_text())
            urlTimes.append(Times2[i]['href'])

        return Times, urlTimes, media_gols_liga

    except HTTPError as e:
        print(e.status, e.reason)

    except URLError as e:
        print(e.reason)

def buscarTime(urlTime):
        Meses = {
            'Jan': '1',
            'Feb': '2',
            'Mar': '3',
            'Apr': '4',
            'May': '5',
            'Jun': '6',
            'Jul': '7',
            'Aug': '8',
            'Sep': '9',
            'Oct': '10',
            'Nov': '11',
            'Dec': '12',
        }
        try:
            #team.asp?league=brazil&stats=2-santos

            liga = urlTime.split('=')[1].split('&')[0]
            time = urlTime.split('&')[1].split('=')[1].split('-')[0]

            # print(liga, time)

            params = {
                'league': liga,
                }
            data = {
                'theteams': time,
                'themonth': '1', #1 puxa jogados em casa 2 fora de casa 99 todos
                'theoutcome': '99',
                'thehometeam': '99',
                'theawayteam': '99',
                'thematchgoals': '99',
                'theHToutcome': '99',
                'theFG': '99',}


            response = requests.post('https://www.soccerstats.com/matchlist.asp', params=params, data=data)

            

            soup = BeautifulSoup(response.content, 'html.parser')


            dataMedia = {
                'theteams': time,
                'themonth': '99', #1 puxa jogados em casa 2 fora de casa 99 todos
                'theoutcome': '99',
                'thehometeam': '99',
                'theawayteam': '99',
                'thematchgoals': '99',
                'theHToutcome': '99',
                'theFG': '99',}

            responseMedia = requests.post('https://www.soccerstats.com/matchlist.asp', params=params, data=dataMedia)
            soupMedia = BeautifulSoup(responseMedia.content, 'html.parser')

            media_gols_time = ""


            for item in soupMedia.find_all("td"):
                if not "Averages" in item.text:continue
                media_gols_time = item.parent.td.next_sibling.next_sibling.next_sibling.next_sibling.next_sibling.next_sibling.next_sibling.text


            soup.tfoot.decompose()
            table_raw = soup.find('td', string='Matches played by:').parent.parent.next_sibling.next_sibling


            table_MN = pd.read_html(str(table_raw))[0]
            df_F1 = table_MN
            #df_F = table_MN.drop(['Date', 'Outcome', '1.5+', '2.5+', '3.5+', 'TG', 'BTS', 'hCS', 'aCS', 'HTh', 'HTa', 'FG', 'Unnamed: 16' ], axis=1)
            df_F1 = df_F1.rename(columns={'Date':'Data','Local': 'Time A', 'Visitor': 'Adversário', 'HG': 'Gols Time A', 'AG': 'Gols Adversário'})
            df_F1 = df_F1[['Data','Time A', 'Gols Time A', 'Adversário', 'Gols Adversário']]

            for i, values in enumerate(df_F1['Data']):
                for meses in Meses:
                    if meses in values:
                        # print(Meses[meses], value)
                        dataaa = values.rsplit(' ', 1)[0]+f'-{Meses[meses]}-2022'
                        dataaa = pd.to_datetime(dataaa, dayfirst=True, format='%d-%m-%Y')
                        df_F1.loc[i, ['Data']] = dataaa 

            data2 = {
                'theteams': time,
                'themonth': '2', #1 puxa jogados em casa 2 fora de casa 99 todos
                'theoutcome': '99',
                'thehometeam': '99',
                'theawayteam': '99',
                'thematchgoals': '99',
                'theHToutcome': '99',
                'theFG': '99',}

            response2 = requests.post('https://www.soccerstats.com/matchlist.asp', params=params, data=data2)
            

            soup2 = BeautifulSoup(response2.content, 'html.parser')
            soup2.tfoot.decompose()
            table_raw2 = soup2.find('td', string='Matches played by:').parent.parent.next_sibling.next_sibling

            table_MN2 = pd.read_html(str(table_raw2))[0]
            df_F2 = table_MN2
            #df_F = table_MN.drop(['Date', 'Outcome', '1.5+', '2.5+', '3.5+', 'TG', 'BTS', 'hCS', 'aCS', 'HTh', 'HTa', 'FG', 'Unnamed: 16' ], axis=1)
            df_F2 = df_F2.rename(columns={'Date':'Data', 'Local': 'Adversário', 'Visitor': 'Time A', 'HG': 'Gols Adversário', 'AG': 'Gols Time A'})
            df_F2 = df_F2[['Data', 'Time A', 'Gols Time A', 'Adversário', 'Gols Adversário']]
            
            for i, values in enumerate(df_F2['Data']):
                for meses in Meses:
                    if meses in values:
                        # print(Meses[meses], value)
                        dataaa = values.rsplit(' ', 1)[0]+f'-{Meses[meses]}-2022'
                        dataaa = pd.to_datetime(dataaa, dayfirst=True, format='%d-%m-%Y')
                        df_F2.loc[i, ['Data']] = dataaa

            df_F = pd.concat([df_F1, df_F2], ignore_index=True).sort_values(by='Data',ascending=False)

            df_F.insert(5, "Total Gols", "#")
            df_F['Data'] = pd.to_datetime(df_F["Data"]).dt.date


            #df_F.to_excel("teste.xlsx", index=False)



            return df_F, media_gols_time

            media_gols = ""
            resultados = {
                'Data': '',
                'Adversário': '',
                'Gol Sofritos': '',
                'Gols Feitos': ''
            }

            # Times = []
            # urlTimes = []
            # Times1 = soup.find('font', string='Teams')
            # Times2 = Times1.parent.parent.find_all('a', href=True)


            # for i, t in enumerate(Times2):
            #     Times.append(Times2[i].get_text())
            #     urlTimes.append(Times2[i]['href'])

        except HTTPError as e:
            print(e.status, e.reason)

        except URLError as e:
            print(e.reason)


def adicionarPlanilha(dataframe_timeA, dataframe_timeB, nome_planilha, media_gols_liga):
    isFile = os.path.isfile(nome_planilha)
    if(isFile == False):
        writer = pd.ExcelWriter(nome_planilha, engine='xlsxwriter')
        writer.close()

    

    df_F = dataframe_timeA[0].reset_index(drop=True).merge(dataframe_timeB[0].reset_index(drop=True), left_index=True, right_index=True)
    
    df_F = df_F.rename(columns={'Data_x': 'Data A','Time A_x': 'Time A', 'Gols Time A_x': 'Gols Time A', 'Adversário_x': 'Adversário Time A', 'Gols Adversário_x': 'Gols Adversário Time A', 'Total Gols_x': 'Total Gols A' ,'Data_y': 'Data B','Time A_y': 'Time B', 'Gols Time A_y': 'Gols Time B', 'Adversário_y': 'Adversário Time B', 'Gols Adversário_y': 'Gols Adversário Time B', 'Total Gols_y': 'Total Gols B' })
    df_F['Total Gols A'] = df_F['Gols Time A'] + df_F['Gols Adversário Time A']
    df_F['Total Gols B'] = df_F['Gols Time B'] + df_F['Gols Adversário Time B']



    try:

        writer = pd.ExcelWriter(nome_planilha, engine='openpyxl', if_sheet_exists='replace', mode='a', date_format='d/m/yyyy')
        df_F.to_excel(writer, sheet_name='Resultado', startrow=7, startcol=2, index=False, na_rep='NaN')



        workbook  = writer.book
        std=workbook.get_sheet_names()
        if(std[0] == "Sheet1"):
            std2=workbook.get_sheet_by_name('Sheet1')
            workbook.remove_sheet(std2)

        sheet = workbook.get_sheet_by_name('Resultado')
        for column_cells in sheet.columns:
            new_column_length = max(len(str(cell.value)) for cell in column_cells)
            new_column_letter = (get_column_letter(column_cells[0].column))
            if new_column_length > 0:
                sheet.column_dimensions[new_column_letter].width = new_column_length*1.23

        for row in range(1,sheet.max_row+1):
            for col in range(1,sheet.max_column+1):
                cell=sheet.cell(row, col)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        
        writer.close()


        mediag = [[dataframe_timeA[0]['Time A'][0], dataframe_timeA[1]], [dataframe_timeB[0]['Time A'][0], dataframe_timeB[1]]]
        dfMedia = pd.DataFrame(mediag, columns=['Time', 'Média de Gols'])
        dfMedia = dfMedia.stack().str.replace('.',',', regex=True).unstack()
        

        writer2 = pd.ExcelWriter(nome_planilha, engine='openpyxl', if_sheet_exists='overlay', mode='a', date_format='d/m/yyyy')
        dfMedia.to_excel(writer2, sheet_name='Resultado', startrow=3 , startcol=0, index=False, na_rep='NaN')
        
        workbook2 = writer2.book
        
        sheet2 = workbook2.active
        sheet2.column_dimensions['A'].width = 20
        sheet2.column_dimensions['B'].width = 20
		
        writer2.close()

        dfMedia3 = pd.DataFrame(media_gols_liga, columns=['Média de Gols Liga'])
        dfMedia3 = dfMedia3.stack().str.replace('.',',', regex=True).unstack()


        writer3 = pd.ExcelWriter(nome_planilha, engine='openpyxl', if_sheet_exists='overlay', mode='a', date_format='d/m/yyyy')
        dfMedia3.to_excel(writer3, sheet_name='Resultado', startrow=0, startcol=0, index=False, na_rep='NaN')
        
		
        writer3.close()

        # df = pd.read_excel(nome_planilha)

        # update_xlsx(nome_planilha, nome_planilha, df)

        return "ok"

    except PermissionError as pe:
        return pe.errno


# print(resultadoo)